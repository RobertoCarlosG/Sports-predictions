"""Exportación de apuestas por periodo a Excel (.xlsx)."""

from __future__ import annotations

import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.bets import Bet, BetPeriod
from app.models.mlb import Game
from app.services.bets_service import bet_realized_profit, _build_period_stats


def _pick_label(bet: Bet) -> str:
    if bet.bet_type == "moneyline":
        side = "Local" if bet.bet_side == "home" else "Visitante"
        return f"Ganador · {side}"
    side = "Más" if bet.bet_side == "over" else "Menos"
    line = bet.ou_line if bet.ou_line is not None else ""
    return f"Más/menos · {side} ({line})"


def _result_label(status: str) -> str:
    return {
        "pending": "Pendiente",
        "won": "Ganada",
        "lost": "Perdida",
        "push": "Empate",
        "cancelled": "Cancelada",
    }.get(status, status)


def _fill_for_pick(status: str) -> PatternFill | None:
    if status == "lost":
        return PatternFill(fill_type="solid", fgColor="FFCCCC")
    if status == "won":
        return PatternFill(fill_type="solid", fgColor="CCFFCC")
    if status == "push":
        return PatternFill(fill_type="solid", fgColor="FFFFCC")
    return None


async def build_period_workbook_bytes(
    period: BetPeriod,
    bets: list[Bet],
    game_labels: dict[int, str],
    game_dates: dict[int, dt.date],
) -> bytes:
    """Genera bytes XLSX; `game_labels` map game_pk -> etiqueta legible (ej. LAD @ NYY)."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Apuestas"

    headers = ["Fecha", "Partido", "Tipo", "Predicción", "Línea", "Apuesta", "Cuota", "Resultado", "P/L"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)

    stats = _build_period_stats(period, bets)
    row_idx = 2
    for b in bets:
        if b.status == "cancelled":
            continue
        gdate = game_dates.get(b.game_pk)
        matchup = game_labels.get(b.game_pk, "Partido")
        tipo = "Ganador" if b.bet_type == "moneyline" else "Más/menos"
        pick = _pick_label(b)
        line_val = b.ou_line if b.bet_type == "over_under" else None
        ws.cell(row=row_idx, column=1, value=gdate.isoformat() if gdate else "")
        ws.cell(row=row_idx, column=2, value=matchup)
        ws.cell(row=row_idx, column=3, value=tipo)
        cell_pick = ws.cell(row=row_idx, column=4, value=pick)
        fl = _fill_for_pick(b.status)
        if fl is not None:
            cell_pick.fill = fl
        ws.cell(row=row_idx, column=5, value=line_val if line_val is not None else "")
        ws.cell(row=row_idx, column=6, value=float(b.stake))
        ws.cell(row=row_idx, column=7, value=float(b.odds))
        ws.cell(row=row_idx, column=8, value=_result_label(b.status))
        pnl = bet_realized_profit(b.stake, b.odds, b.status) if b.status in ("won", "lost", "push") else None
        ws.cell(row=row_idx, column=9, value=pnl if pnl is not None else "")
        row_idx += 1

    ws2 = wb.create_sheet("Resumen")
    ws2.cell(row=1, column=1, value="Periodo")
    ws2.cell(row=1, column=2, value=period.name)
    ws2.cell(row=2, column=1, value="Saldo inicial")
    ws2.cell(row=2, column=2, value=stats.starting_balance)
    ws2.cell(row=3, column=1, value="P/L realizado")
    ws2.cell(row=3, column=2, value=stats.realized_pnl)
    ws2.cell(row=4, column=1, value="Apuestas decididas")
    ws2.cell(row=4, column=2, value=stats.decided_bets)
    ws2.cell(row=5, column=1, value="Ganadas / Perdidas / Push")
    ws2.cell(row=5, column=2, value=f"{stats.wins} / {stats.losses} / {stats.pushes}")
    if stats.roi_pct is not None:
        ws2.cell(row=6, column=1, value="ROI %")
        ws2.cell(row=6, column=2, value=round(stats.roi_pct, 2))

    for sheet in wb.worksheets:
        for letter in "ABCDEFGHIJK":
            sheet.column_dimensions[letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def load_game_labels(session: AsyncSession, game_pks: list[int]) -> tuple[dict[int, str], dict[int, dt.date]]:
    if not game_pks:
        return {}, {}
    r = await session.execute(
        select(Game)
        .where(Game.game_pk.in_(game_pks))
        .options(selectinload(Game.home_team), selectinload(Game.away_team)),
    )
    games = list(r.scalars().unique().all())
    labels: dict[int, str] = {}
    dates: dict[int, dt.date] = {}
    for g in games:
        ha = getattr(g, "home_team", None)
        aa = getattr(g, "away_team", None)
        hab = ha.abbreviation if ha else "?"
        aab = aa.abbreviation if aa else "?"
        labels[g.game_pk] = f"{aab} @ {hab}"
        dates[g.game_pk] = g.game_date
    return labels, dates
